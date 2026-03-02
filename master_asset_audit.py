#!/usr/bin/env python3
"""
Master Asset Audit Script
Cross-references assets/, media.js, and chapters.js
"""

import os
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Paths
BASE_DIR = Path("/Users/fenech/off-the-record")
ASSETS_DIR = BASE_DIR / "assets"
MEDIA_JS = BASE_DIR / "data/media.js"
CHAPTERS_JS = BASE_DIR / "data/chapters.js"
OUTPUT_FILE = BASE_DIR / "master-asset-audit.xlsx"

def get_file_type(filename):
    """Determine file type from extension"""
    ext = filename.lower().split('.')[-1] if '.' in filename else ''
    type_map = {
        'webp': 'photo', 'jpg': 'photo', 'jpeg': 'photo', 'png': 'photo', 'gif': 'photo',
        'mp4': 'video', 'mov': 'video', 'avi': 'video', 'webm': 'video',
        'mp3': 'audio', 'wav': 'audio', 'm4a': 'audio', 'ogg': 'audio',
        'pdf': 'document', 'txt': 'document', 'doc': 'document', 'docx': 'document',
        'gitkeep': 'system'
    }
    return type_map.get(ext, 'unknown')

def get_chapter_from_path(filepath):
    """Extract chapter number from file path"""
    path_str = str(filepath)
    # Match patterns like "04-scandal", "14-prime-time", etc.
    match = re.search(r'/(\d{2})-[^/]+/', path_str)
    if match:
        return int(match.group(1))
    # Check for audio-staging or root assets
    if 'audio-staging' in path_str:
        return 'staging'
    if 'mentions' in path_str:
        return 'mentions'
    if 'icons' in path_str:
        return 'icons'
    return 'root'

def inventory_assets():
    """Step 1: Inventory all files in assets/"""
    assets = []
    for filepath in ASSETS_DIR.rglob('*'):
        if filepath.is_file():
            rel_path = filepath.relative_to(BASE_DIR)
            assets.append({
                'filename': filepath.name,
                'path': str(rel_path),
                'type': get_file_type(filepath.name),
                'size': filepath.stat().st_size,
                'chapter': get_chapter_from_path(filepath)
            })
    return assets

def parse_media_js():
    """Step 2: Parse media.js for all entries"""
    content = MEDIA_JS.read_text(encoding='utf-8')
    entries = []

    # Pattern to match media entries
    entry_pattern = r"'([^']+)':\s*\{([^}]+(?:\{[^}]*\}[^}]*)*)\}"

    for match in re.finditer(entry_pattern, content):
        media_id = match.group(1)
        entry_content = match.group(2)

        # Extract fields (handle both single and double quotes)
        type_match = re.search(r"type:\s*['\"]([^'\"]+)['\"]", entry_content)
        emoji_match = re.search(r"emoji:\s*['\"]([^'\"]+)['\"]", entry_content)
        src_match = re.search(r"src:\s*['\"]([^'\"]+)['\"]", entry_content)
        chapter_match = re.search(r"chapter:\s*(\d+)", entry_content)

        media_type = type_match.group(1) if type_match else 'unknown'
        emoji = emoji_match.group(1) if emoji_match else ''
        src = src_match.group(1) if src_match else ''
        chapter = int(chapter_match.group(1)) if chapter_match else None

        # Check if src is a placeholder (external URL or missing)
        is_placeholder = src.startswith('http') or src == ''

        # Check if file exists
        if src and not src.startswith('http'):
            file_exists = (BASE_DIR / src).exists()
        else:
            file_exists = False if not src.startswith('http') else 'N/A (URL)'

        entries.append({
            'media_id': media_id,
            'type': media_type,
            'emoji': emoji,
            'src': src,
            'chapter': chapter,
            'is_placeholder': is_placeholder,
            'file_exists': file_exists
        })

    return entries

def parse_chapters_js():
    """Step 3: Parse chapters.js for all media-emoji spans"""
    content = CHAPTERS_JS.read_text(encoding='utf-8')
    emojis = []

    # Pattern for media-emoji spans
    pattern = r'<span\s+class="media-emoji"\s+data-media-id="([^"]+)"[^>]*>([^<]+)</span>'

    lines = content.split('\n')
    current_chapter = None

    for i, line in enumerate(lines, 1):
        # Track chapter
        id_match = re.search(r"id:\s*(\d+)", line)
        if id_match:
            current_chapter = int(id_match.group(1))

        # Find emoji spans
        for match in re.finditer(pattern, line):
            media_id = match.group(1)
            emoji = match.group(2).strip()

            emojis.append({
                'media_id': media_id,
                'emoji': emoji,
                'chapter': current_chapter,
                'line': i
            })

    return emojis

def format_size(size_bytes):
    """Format file size in human-readable format"""
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    else:
        return f"{size_bytes / (1024 * 1024):.1f} MB"

def format_worksheet(ws, headers):
    """Format worksheet with headers"""
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'

def auto_fit_columns(ws, max_width=60):
    """Auto-fit column widths"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, max_width)

def main():
    print("=" * 60)
    print("MASTER ASSET AUDIT")
    print("=" * 60)

    # Step 1: Inventory assets
    print("\nStep 1: Inventorying assets/...")
    assets = inventory_assets()
    # Filter out system files
    assets = [a for a in assets if a['type'] != 'system']
    print(f"  Found {len(assets)} files")

    # Step 2: Parse media.js
    print("\nStep 2: Parsing media.js...")
    media_entries = parse_media_js()
    print(f"  Found {len(media_entries)} entries")

    # Step 3: Parse chapters.js
    print("\nStep 3: Parsing chapters.js emojis...")
    chapter_emojis = parse_chapters_js()
    print(f"  Found {len(chapter_emojis)} emojis")

    # Build lookup tables
    media_by_src = {e['src']: e for e in media_entries if e['src']}
    media_by_id = {e['media_id']: e for e in media_entries}
    emoji_by_id = {e['media_id']: e for e in chapter_emojis}

    # Create workbook
    wb = Workbook()

    # ========================================
    # SHEET 1: All Assets
    # ========================================
    print("\nBuilding Sheet 1: All Assets...")
    ws1 = wb.active
    ws1.title = "All Assets"

    headers = ['Filename', 'Path', 'Type', 'Size', 'In media.js?', 'In chapters.js?', 'Chapter']
    format_worksheet(ws1, headers)

    for row_num, asset in enumerate(sorted(assets, key=lambda x: x['path']), 2):
        # Check if in media.js
        in_media = 'assets/' + asset['path'].replace('assets/', '') in media_by_src or \
                   asset['path'] in media_by_src

        # Find media_id if in media.js
        media_id = None
        for src, entry in media_by_src.items():
            if asset['path'] in src or src in asset['path']:
                media_id = entry['media_id']
                break

        # Check if in chapters.js
        in_chapters = media_id in emoji_by_id if media_id else False

        ws1.cell(row=row_num, column=1, value=asset['filename'])
        ws1.cell(row=row_num, column=2, value=asset['path'])
        ws1.cell(row=row_num, column=3, value=asset['type'])
        ws1.cell(row=row_num, column=4, value=format_size(asset['size']))
        ws1.cell(row=row_num, column=5, value='Yes' if in_media else 'No')
        ws1.cell(row=row_num, column=6, value='Yes' if in_chapters else 'No')
        ws1.cell(row=row_num, column=7, value=str(asset['chapter']))

    auto_fit_columns(ws1)

    # ========================================
    # SHEET 2: Placed in Manuscript
    # ========================================
    print("Building Sheet 2: Placed in Manuscript...")
    ws2 = wb.create_sheet("Placed in Manuscript")

    headers = ['Media ID', 'Emoji', 'Chapter', 'File Path', 'File Exists?', 'Status']
    format_worksheet(ws2, headers)

    row_num = 2
    for emoji_entry in sorted(chapter_emojis, key=lambda x: (x['chapter'] or 0, x['media_id'])):
        media_id = emoji_entry['media_id']
        media_entry = media_by_id.get(media_id, {})

        src = media_entry.get('src', 'NOT IN MEDIA.JS')

        # Determine file existence
        if src.startswith('http'):
            file_exists = 'N/A (URL)'
            status = 'External Link'
        elif src and src != 'NOT IN MEDIA.JS':
            exists = (BASE_DIR / src).exists()
            file_exists = 'Yes' if exists else 'No'
            status = 'OK' if exists else 'FILE MISSING'
        else:
            file_exists = 'No'
            status = 'NOT IN MEDIA.JS'

        ws2.cell(row=row_num, column=1, value=media_id)
        ws2.cell(row=row_num, column=2, value=emoji_entry['emoji'])
        ws2.cell(row=row_num, column=3, value=emoji_entry['chapter'])
        ws2.cell(row=row_num, column=4, value=src)
        ws2.cell(row=row_num, column=5, value=file_exists)
        ws2.cell(row=row_num, column=6, value=status)
        row_num += 1

    auto_fit_columns(ws2)

    # ========================================
    # SHEET 3: Uploaded but NOT Placed
    # ========================================
    print("Building Sheet 3: Uploaded but NOT Placed...")
    ws3 = wb.create_sheet("Uploaded NOT Placed")

    headers = ['Filename', 'Path', 'Type', 'Size', 'In media.js?', 'Suggested Chapter']
    format_worksheet(ws3, headers)

    row_num = 2
    for asset in sorted(assets, key=lambda x: (str(x['chapter']), x['path'])):
        # Check if this asset has an emoji in chapters.js
        asset_path = asset['path']

        # Find if any media entry points to this file
        media_id = None
        for src, entry in media_by_src.items():
            # Normalize paths for comparison
            if asset_path.replace('assets/', '') in src or src.replace('assets/', '') in asset_path:
                media_id = entry['media_id']
                break

        # If not in chapters.js, it's unplaced
        if not media_id or media_id not in emoji_by_id:
            in_media = media_id is not None

            ws3.cell(row=row_num, column=1, value=asset['filename'])
            ws3.cell(row=row_num, column=2, value=asset['path'])
            ws3.cell(row=row_num, column=3, value=asset['type'])
            ws3.cell(row=row_num, column=4, value=format_size(asset['size']))
            ws3.cell(row=row_num, column=5, value='Yes' if in_media else 'No')
            ws3.cell(row=row_num, column=6, value=str(asset['chapter']))
            row_num += 1

    auto_fit_columns(ws3)

    # ========================================
    # SHEET 4: In Manuscript but Missing File
    # ========================================
    print("Building Sheet 4: In Manuscript but Missing File...")
    ws4 = wb.create_sheet("Missing Files")

    headers = ['Media ID', 'Emoji', 'Chapter', 'Expected Path', 'Status']
    format_worksheet(ws4, headers)

    row_num = 2
    for emoji_entry in sorted(chapter_emojis, key=lambda x: (x['chapter'] or 0, x['media_id'])):
        media_id = emoji_entry['media_id']
        media_entry = media_by_id.get(media_id, {})

        src = media_entry.get('src', '')

        # Check for missing or placeholder
        is_missing = False
        status = ''

        if not media_entry:
            is_missing = True
            status = 'NOT IN MEDIA.JS'
            src = 'N/A'
        elif src.startswith('http'):
            # External URL - check if it's a placeholder URL
            if src in ['https://www.freep.com/', 'https://www.espn.com/',
                       'https://www.latimes.com/', 'https://www.theringer.com/',
                       'https://www.si.com/', 'https://www.baseball-reference.com/',
                       'https://www.mtv.com/']:
                is_missing = True
                status = 'PLACEHOLDER URL'
        elif src:
            if not (BASE_DIR / src).exists():
                is_missing = True
                status = 'FILE MISSING'
        else:
            is_missing = True
            status = 'NO SRC PATH'

        if is_missing:
            ws4.cell(row=row_num, column=1, value=media_id)
            ws4.cell(row=row_num, column=2, value=emoji_entry['emoji'])
            ws4.cell(row=row_num, column=3, value=emoji_entry['chapter'])
            ws4.cell(row=row_num, column=4, value=src)
            ws4.cell(row=row_num, column=5, value=status)
            row_num += 1

    auto_fit_columns(ws4)

    # ========================================
    # SHEET 5: Summary
    # ========================================
    print("Building Sheet 5: Summary...")
    ws5 = wb.create_sheet("Summary")

    # Calculate stats
    total_assets = len(assets)
    total_media_entries = len(media_entries)
    total_emojis = len(chapter_emojis)

    # Files uploaded but not placed
    placed_ids = set(emoji_by_id.keys())
    unplaced_count = 0
    for asset in assets:
        asset_path = asset['path']
        media_id = None
        for src, entry in media_by_src.items():
            if asset_path.replace('assets/', '') in src or src.replace('assets/', '') in asset_path:
                media_id = entry['media_id']
                break
        if not media_id or media_id not in placed_ids:
            unplaced_count += 1

    # Emojis with missing files
    missing_count = 0
    for emoji_entry in chapter_emojis:
        media_id = emoji_entry['media_id']
        media_entry = media_by_id.get(media_id, {})
        src = media_entry.get('src', '')

        if not media_entry:
            missing_count += 1
        elif src.startswith('http'):
            if src in ['https://www.freep.com/', 'https://www.espn.com/',
                       'https://www.latimes.com/', 'https://www.theringer.com/',
                       'https://www.si.com/', 'https://www.baseball-reference.com/',
                       'https://www.mtv.com/']:
                missing_count += 1
        elif src and not (BASE_DIR / src).exists():
            missing_count += 1
        elif not src:
            missing_count += 1

    # Fully wired up
    fully_wired = total_emojis - missing_count

    # Write summary
    summary_data = [
        ('MASTER ASSET AUDIT SUMMARY', ''),
        ('', ''),
        ('Total files in assets/', total_assets),
        ('Total entries in media.js', total_media_entries),
        ('Total emojis in chapters.js', total_emojis),
        ('', ''),
        ('Files uploaded but NOT placed', unplaced_count),
        ('Emojis with missing/placeholder files', missing_count),
        ('Fully wired up (file + media.js + chapters.js)', fully_wired),
        ('', ''),
        ('Coverage Rate', f"{(fully_wired / total_emojis * 100):.1f}%" if total_emojis > 0 else "N/A"),
    ]

    for row_num, (label, value) in enumerate(summary_data, 1):
        cell = ws5.cell(row=row_num, column=1, value=label)
        if row_num == 1:
            cell.font = Font(bold=True, size=14)
        elif label and not value == '':
            cell.font = Font(bold=True)
        ws5.cell(row=row_num, column=2, value=value)

    ws5.column_dimensions['A'].width = 45
    ws5.column_dimensions['B'].width = 15

    # Save
    wb.save(OUTPUT_FILE)
    print(f"\nSaved to: {OUTPUT_FILE}")

    # Print summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Total files in assets/: {total_assets}")
    print(f"Total entries in media.js: {total_media_entries}")
    print(f"Total emojis in chapters.js: {total_emojis}")
    print(f"Files uploaded but NOT placed: {unplaced_count}")
    print(f"Emojis with missing files: {missing_count}")
    print(f"Fully wired up: {fully_wired}")
    print(f"Coverage Rate: {(fully_wired / total_emojis * 100):.1f}%" if total_emojis > 0 else "N/A")

if __name__ == "__main__":
    main()
